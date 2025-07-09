import time
from datetime import date
import openpyxl

from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

from pythonProject.share_price.Common_share_data import *

# import undetected_chromedriver as uc

# Replace with your proxy
proxy = "10.20.2.60"

ops = webdriver.ChromeOptions()
ops.add_argument("--disable-notifications")
# ops.add_argument('--headless')
#ops.add_argument(f'--proxy-server=http://{proxy}')

workbook = openpyxl.load_workbook(path_File)
sheet = workbook.active
data_list = []
share_price = []

'''Get the Data from the Excel'''

for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming data starts from the second row
    data_dict = {
        'stock Name': row[0]
    }
    data_list.append(data_dict)
workbook.close()

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=ops)
driver.maximize_window()
driver.implicitly_wait(10)
driver.get(URL)
driver.find_element(By.XPATH, "//textarea[@name='q']").send_keys("hello", Keys.ENTER)
time.sleep(30)

'''Get the Data from the Google'''
for data in data_list:
    search_key = driver.find_element(By.XPATH, "//textarea[@name='q']")
    value = data['stock Name']
    search_key.clear()
    search_key.send_keys(value, Keys.ENTER)
    price_value = driver.find_element(By.XPATH, "//div[@class='PZPZlf']/span[1]/span/span[1]").text
    print(price_value)
    share_price.append(price_value)

driver.quit()

today = date.today()
workbook = openpyxl.load_workbook(path_File)
sheet = workbook.active
sheet.insert_cols(2)
sheet.cell(row=1, column=2, value=today)

for i, value in enumerate(share_price, start=1):
    sheet.cell(row=i + 1, column=2, value=value)

workbook.save(path_File)

# time.sleep(20)
