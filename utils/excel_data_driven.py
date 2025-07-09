import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import openpyxl


def getRowCount(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def getColumnCount(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def readData(file_path, sheet_name, rownum, columnnum):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file_path, sheet_name, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=rownum, column=columnnum, value=data)
    workbook.save(file_path)


def get_field_names(file_path, sheet_name):
    """
    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read from.
    :return: List of dictionaries containing stock names.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    field_list = []

    # Read data starting from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_dict = {'Field Name': row[0]}  # Assuming stock name is in the first column
        field_list.append(data_dict)

    workbook.close()
    return field_list  # Output: [{'Field Name': 'ITC'}, {'Field Name': 'HDFC'}, {'Field Name': 'HUL'}]
